import logging
import os
import openpyxl
import asyncio
from telegram import (
    Update, KeyboardButton, ReplyKeyboardMarkup,
    ReplyKeyboardRemove
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    filters, ContextTypes, ConversationHandler
)

# Logger
logging.basicConfig(level=logging.INFO)

# Fayl nomi
FILE_NAME = "users.xlsx"

# Admin Telegram ID
ADMIN_ID = 1350513135

# Bosqichlar
COURSE, PHONE, AGE, DAY, TIME, CONFIRM_DAY = range(6)

# Kurslar ro'yxati
COURSES = [
    "🟢 START POINT", "🤖 ROBOTICS", "⚙️ CHALLENGE LAB",
    "✈️ FLIGHT ACADEMY", "🧪 SCINECE LAB", "🏗️ ENGINEERING LAB",
    "💻 CODING ROOM", "🎮 VR ROOM", "🔧 VEX V5- IQ ROOM"
]

# Kunlar
DAYS = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]

# Soatlar
TIMES = ["09:00-11:00", "11:00-13:00", "14:00-16:00", "16:00-18:00"]

# Excel faylni tayyorlash
def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Ism", "Telefon", "Yosh", "Kurs", "Kunlar", "Vaqt"])
        wb.save(FILE_NAME)

# Xabarlarni o'chirish funksiyasi
async def delete_messages(context: ContextTypes.DEFAULT_TYPE, chat_id: int, message_ids: list, delay: int = 10):
    await asyncio.sleep(delay)
    for msg_id in message_ids:
        try:
            await context.bot.delete_message(chat_id=chat_id, message_id=msg_id)
        except:
            pass

# /start komandasi
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    init_excel()
    welcome_text = (
        "👋 *Assalomu alaykum!*\n\n"
        "Sizga _\"Abdulla Avloniy nomidagi Pedagogik Mahorat Milliy Instituti\"_ STEAM markazi tomonidan tashkil etilgan "
        "*innovatsion kurslar* bo‘yicha ro‘yxatdan o‘tish uchun bir nechta savollar beriladi.\n\n"
        "📌 *Bu markaz* zamonaviy laboratoriyalar, ilg‘or texnologiyalar va amaliy loyihalar asosida ta’lim beradi. "
        "Har bir yo‘nalish o‘quvchilarning bilim olishiga, *ixtirochilik salohiyatini* oshirishga qaratilgan.\n\n"
        "🧭 *Iltimos, quyidagi yo‘nalishlardan birini tanlang:*"
    )

    keyboard = [[KeyboardButton(course)] for course in COURSES]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    user_msg = update.message.message_id
    bot_msg = await update.message.reply_text(welcome_text, parse_mode="Markdown", reply_markup=reply_markup)
    await delete_messages(context, update.effective_chat.id, [user_msg, bot_msg.message_id], delay=10)
    return COURSE

# Kurs tanlash
async def course_chosen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text not in COURSES:
        msg = await update.message.reply_text("Iltimos, menyudan kursni tanlang.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return COURSE
    context.user_data["course"] = text
    contact_button = KeyboardButton("📞 Telefon raqamni yuborish", request_contact=True)
    reply_markup = ReplyKeyboardMarkup([[contact_button]], resize_keyboard=True, one_time_keyboard=True)
    msg = await update.message.reply_text("Telefon raqamingizni yuboring:", reply_markup=reply_markup)
    await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id], delay=10)
    return PHONE

# Telefon raqam qabul qilish
async def phone_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact = update.message.contact
    if not contact:
        msg = await update.message.reply_text("Iltimos, telefon raqam tugmasidan foydalaning.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return PHONE
    context.user_data["phone"] = contact.phone_number
    context.user_data["name"] = update.effective_user.full_name
    msg = await update.message.reply_text("Yoshingizni kiriting:", reply_markup=ReplyKeyboardRemove())
    await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
    return AGE

# Yosh
async def age_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    age = update.message.text
    if not age.isdigit():
        msg = await update.message.reply_text("Faqat raqam kiriting. Yoshingizni qayta kiriting:")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return AGE
    context.user_data["age"] = age
    context.user_data["days"] = []
    return await send_day_keyboard(update, context)

# Kunlar uchun klaviatura
async def send_day_keyboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    selected = context.user_data.get("days", [])
    keyboard = []
    for day in DAYS:
        display = f"✅ {day}" if day in selected else day
        keyboard.append([KeyboardButton(display)])
    keyboard.append([KeyboardButton("✅ Tayyor")])
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    msg = await update.message.reply_text("Qaysi kunlarda qatnasha olasiz? Bir nechta tanlang:", reply_markup=reply_markup)
    await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
    return CONFIRM_DAY

# Kun tanlash
async def day_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.replace("✅", "").strip()
    if text == "Tayyor":
        if not context.user_data["days"]:
            msg = await update.message.reply_text("Iltimos, kamida bitta kun tanlang.")
            await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
            return await send_day_keyboard(update, context)
        return await ask_time(update, context)
    if text in DAYS:
        selected = context.user_data["days"]
        if text in selected:
            selected.remove(text)
        else:
            if len(selected) < 7:
                selected.append(text)
            else:
                msg = await update.message.reply_text("Eng ko‘pi bilan 7 ta kun tanlashingiz mumkin.")
                await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return await send_day_keyboard(update, context)
    else:
        msg = await update.message.reply_text("Iltimos, menyudan tanlang.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return await send_day_keyboard(update, context)

# Soat tanlash
async def ask_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[KeyboardButton(time)] for time in TIMES]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
    msg = await update.message.reply_text("Qaysi soatlarda qatnasha olasiz?", reply_markup=reply_markup)
    await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
    return TIME

# Soat qabul qilish va ma'lumotlarni saqlash
async def time_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    time = update.message.text
    if time not in TIMES:
        msg = await update.message.reply_text("Iltimos, soatlardan birini tanlang.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return TIME
    context.user_data["time"] = time
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([
        context.user_data["name"],
        context.user_data["phone"],
        context.user_data["age"],
        context.user_data["course"],
        ", ".join(context.user_data["days"]),
        context.user_data["time"]
    ])
    wb.save(FILE_NAME)
    msg = await update.message.reply_text("Ma’lumotlar saqlandi. Agar yana boshqa yo'nalishni tanlamoqchi bo'lsangiz /start yozuvi ustiga bosing!", reply_markup=ReplyKeyboardRemove())
    await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id], delay=10)
    return ConversationHandler.END

# Bekor qilish
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("Bekor qilindi.", reply_markup=ReplyKeyboardRemove())
    await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id], delay=10)
    return ConversationHandler.END

# Admin: Excel faylni yuborish
async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        msg = await update.message.reply_text("Siz admin emassiz.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return
    await update.message.reply_document(open(FILE_NAME, "rb"))

# Admin: Matn ko‘rinishida foydalanuvchilar
async def list_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        msg = await update.message.reply_text("Siz admin emassiz.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    text = ""
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        text += f"{i}. {row[0]} - {row[3]} - {row[4]} - {row[5]}\n"
    await update.message.reply_text(text or "Foydalanuvchilar mavjud emas.")

# Admin: Tozalash
async def clear_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        msg = await update.message.reply_text("Siz admin emassiz.")
        await delete_messages(context, update.effective_chat.id, [update.message.message_id, msg.message_id])
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ism", "Telefon", "Yosh", "Kurs", "Kunlar", "Vaqt"])
    wb.save(FILE_NAME)
    await update.message.reply_text("Barcha ma’lumotlar o‘chirildi.")

# Botni ishga tushurish
def main():
    app = ApplicationBuilder().token("YOUR_BOT_TOKEN").build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            COURSE: [MessageHandler(filters.TEXT & ~filters.COMMAND, course_chosen)],
            PHONE: [MessageHandler(filters.CONTACT, phone_received)],
            AGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, age_received)],
            CONFIRM_DAY: [MessageHandler(filters.TEXT & ~filters.COMMAND, day_selection)],
            TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, time_received)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv_handler)
    app.add_handler(CommandHandler("file", export))
    app.add_handler(CommandHandler("list", list_users))
    app.add_handler(CommandHandler("clear", clear_data))
    app.run_polling()

if __name__ == "__main__":
    main()
